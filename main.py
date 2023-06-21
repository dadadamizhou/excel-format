from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def recursion_format_excel(sheet: Worksheet, data_list: list, parent_id: int = 0, row: int = 0,
                           column: int = 0) -> list:
    column += 1
    res = {
        "row": row,
        "column": column
    }
    data_index = 1
    for d in data_list:
        if d["parent_id"] == parent_id:
            if data_index > 1 or res["row"] == 0:
                res["row"] += 1
            data_index += 1
            recursion_res = recursion_format_excel(sheet=sheet, data_list=data_list, parent_id=d["data_id"],
                                                   row=res["row"], column=column)
            if recursion_res["row"] == res["row"]:
                cell = sheet.cell(row=res["row"], column=column, value=d["name"])
            else:
                column_letter = get_column_letter(column)
                merge_range = f"{column_letter}{res['row']}:{column_letter}{recursion_res['row']}"
                sheet.merge_cells(merge_range)
                cell = sheet[f"{column_letter}{res['row']}"]
                cell.value = d["name"]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if d["bgc"].rgb != '00000000':
                cell.fill = PatternFill(fill_type='solid', fgColor=d["bgc"])
            res = recursion_res
    return res


def format_excel(file_name: str):
    try:
        excel_file = load_workbook(filename=file_name)
        sheet_list = excel_file.sheetnames
        sheet = excel_file[sheet_list[0]]
        data_list = []
        for row in sheet.iter_rows(min_row=2):
            if row[0].value is not None:
                data_list.append({
                    "data_id": row[0].value,
                    "name": row[1].value,
                    "money": row[2].value,
                    "parent_id": row[3].value,
                    "level": row[4].value,
                    "bgc": row[4].fill.fgColor
                })
        excel_file.remove(excel_file[sheet_list[1]])
        sheet2 = excel_file.create_sheet(title='Sheet2')
        recursion_format_excel(sheet=sheet2, data_list=data_list)
        excel_file.save(file_name)
        print("Format Success！！！")
    except Exception as e:
        print("Error:" + str(e))


if __name__ == '__main__':
    default_name = './excel.xlsx'
    input_name = input("请输入excel文件路径(默认`" + default_name + "`):")
    if input_name == "":
        input_name = default_name
    format_excel(file_name=input_name)
